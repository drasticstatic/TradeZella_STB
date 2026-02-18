#!/usr/bin/env python3
"""
TradeZella → SmartTraderAI (STB) Bulk Import Merger
====================================================
Converts a TradeZella CSV export into the STB Bulk Import format.
By default writes directly to Google Sheets. Falls back to .xlsx if no
Google credentials are configured.

USAGE
─────
  # Auto-mode: Google Sheets if credentials exist, otherwise .xlsx
  python3 tradezella_to_stb.py trades_export.csv

  # Force Google Sheets
  python3 tradezella_to_stb.py trades_export.csv --sheets

  # Force .xlsx output
  python3 tradezella_to_stb.py trades_export.csv --xlsx

  # Specify custom template or output path (xlsx mode)
  python3 tradezella_to_stb.py trades_export.csv --xlsx --template path/to/template.xlsx --output out.xlsx

REQUIREMENTS
────────────
  pip install pandas openpyxl gspread google-auth

GOOGLE SHEETS SETUP
───────────────────
  See README.md for full step-by-step instructions.
  Short version:
    1. Create a Google Cloud project, enable Sheets API
    2. Create a Service Account, download JSON key
    3. Share your STB Google Sheet with the service account email
    4. Set SPREADSHEET_ID and SERVICE_ACCOUNT_FILE below (or use --sheet-id / --creds flags)
"""

import sys
import os
import argparse
import pandas as pd
from datetime import datetime

# ─── USER CONFIGURATION ───────────────────────────────────────────────────────
# Update these two values after completing Google Cloud setup (see README).
# You can also pass them as command-line flags instead.

SPREADSHEET_ID       = "YOUR_SPREADSHEET_ID_HERE"
SERVICE_ACCOUNT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "service_account.json")
SHEET_TAB_NAME       = "Sheet1"   # The tab name inside your Google Sheet

# Path to STB template (used for .xlsx fallback mode only)
TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "STB_Import_Template.xlsx")
# ──────────────────────────────────────────────────────────────────────────────


# Valid STB Entry Model dropdown values (lowercase for matching)
VALID_ENTRY_MODELS = {
    '3x entry', 'advanced structure entry', 'breakers',
    'catching the move of the day', 'catching the move of the week',
    'change of delivery', 'cisd', 'displacement', 'fail flip', 'inversions',
    'fcr', 'market structure shift', 'inverted fvg', 'mmem', 'ny fx entry',
    'smm entry', 'other (specify below)', 'time based entry model 2',
    'time based entry model 1'
}


# ─── Field Transformers ───────────────────────────────────────────────────────

def get_entry_model(val):
    """
    Returns (entry_model, other_specify) tuple.
    - Single valid known model → (model, '-')
    - Multiple models tagged → STB only accepts one; use first match,
      put remainder in other_specify column for reference
    - Full dropdown dump (all models present = nothing selected) → ('other (specify below)', '-')
    - Blank → ('other (specify below)', '-')
    Note: normalises legacy 'csid' typo to 'cisd' automatically.
    """
    if pd.isna(val) or str(val).strip() == '':
        return 'other (specify below)', '-'
    # Normalise legacy typo from older TradeZella exports
    normalised = str(val).replace('csid', 'cisd')
    parts = [p.strip().lower() for p in normalised.split(',')]
    matches = [p for p in parts if p in VALID_ENTRY_MODELS]
    non_other = [m for m in matches if m != 'other (specify below)']
    # Full dropdown dump — nothing was actually selected
    all_valid_non_other = VALID_ENTRY_MODELS - {'other (specify below)'}
    if set(non_other) >= all_valid_non_other:
        return 'other (specify below)', '-'
    if non_other:
        # Use first match as the entry model; if multiple, note extras in other column
        extra = ', '.join(non_other[1:]) if len(non_other) > 1 else '-'
        return non_other[0], extra
    return 'other (specify below)', '-'


def get_outcome(row):
    """Map Status + Net P&L → green / red / breakeven."""
    status = str(row.get('Status', '')).strip().lower()
    pnl = row.get('Net P&L', 0)
    try:
        pnl = float(pnl)
    except (TypeError, ValueError):
        pnl = 0
    if status == 'breakeven' or pnl == 0:
        return 'breakeven'
    if status == 'win' or pnl > 0:
        return 'green'
    if status == 'loss' or pnl < 0:
        return 'red'
    return ''


def normalize_yesno(val):
    """
    Convert TradeZella yes/no fields to 'yes' or 'no'.
    If both options are present (e.g. 'yes, no') it means TradeZella exported
    the full dropdown — nothing was actually selected — so return blank.
    """
    v = str(val).strip().lower()
    # Both options present = full dropdown dump, not a real answer
    if 'yes' in v and 'no' in v:
        return ''
    if v in ('true', 'yes', '1', 'y'):
        return 'yes'
    if v in ('false', 'no', '0', 'n'):
        return 'no'
    return ''


def safe_str(val):
    """Return stripped string or empty string for null/nan."""
    if pd.isna(val):
        return ''
    s = str(val).strip()
    return '' if s.lower() == 'nan' else s


def safe_date(val):
    """Parse date safely, return date object or None."""
    try:
        return pd.to_datetime(val).date() if pd.notna(val) else None
    except Exception:
        return None


def format_date(val):
    """Format date as YYYY-MM-DD string — accepted by both Google Sheets and xlsx."""
    d = safe_date(val)
    return d.strftime('%Y-%m-%d') if d else ''


# ─── Row Mapper ───────────────────────────────────────────────────────────────

def map_row(row):
    """Map a single TradeZella row dict to an ordered list of STB column values."""
    entry_model, other_specify = get_entry_model(row.get('Entry Model'))
    return [
        format_date(row.get('Open Date')),                             # A: Trading Date
        entry_model,                                                   # B: Entry Model
        other_specify,                                                 # C: <--Other (Specify)
        'USD',                                                         # D: Currency
        row.get('Net P&L', 0),                                         # E: Profit / Loss
        get_outcome(row),                                              # F: Outcome
        safe_str(row.get('Emotions')),                                 # G: Emotions
        normalize_yesno(row.get('Did Emotions Affect Decisions?')),    # H: Did emotions affect decisions?
        normalize_yesno(row.get('Was Emotionally Stable?')),           # I: Was emotionally stable?
        safe_str(row.get('Profit Target   Did You Respect It?')),      # J: Profit target
        safe_str(row.get('Stop Loss   Did You Respect It?')),          # K: Stop loss
        safe_str(row.get('Entry Logic Explanation')),                  # L: Entry logic explanation
        safe_str(row.get('How Did The Trade Play Out?')),              # M: How did trade play out?
        safe_str(row.get('Notes For Coaches')),                        # N: Notes for coaches
        '',                                                            # O: Screenshot URLs
    ]


# ─── Google Sheets Output ─────────────────────────────────────────────────────

def write_to_sheets(df, spreadsheet_id, service_account_file, tab_name):
    """Append mapped trade rows to the Google Sheet."""
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        print("❌ Missing packages. Run:  pip install gspread google-auth")
        sys.exit(1)

    print(f"🔑 Authenticating with Google...")
    scopes = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive',
    ]
    creds  = Credentials.from_service_account_file(service_account_file, scopes=scopes)
    client = gspread.authorize(creds)

    print(f"📊 Opening spreadsheet...")
    sheet = client.open_by_key(spreadsheet_id).worksheet(tab_name)

    # Find first empty row after existing data
    existing  = sheet.get_all_values()
    next_row  = len(existing) + 1
    data_rows = max(0, len(existing) - 1)
    print(f"   Sheet currently has {data_rows} data row(s). Appending from row {next_row}.")

    rows = [map_row(row) for _, row in df.iterrows()]

    if rows:
        start_cell = f"A{next_row}"
        sheet.update(start_cell, rows, value_input_option='USER_ENTERED')
        print(f"✅ {len(rows)} trades appended → tab '{tab_name}'")
    else:
        print("⚠️  No rows to write.")


# ─── Excel Output ─────────────────────────────────────────────────────────────

def write_to_xlsx(df, template_path, output_path):
    """Write mapped trades into a copy of the STB .xlsx template."""
    from openpyxl import load_workbook

    print(f"📋 Loading STB template: {os.path.basename(template_path)}")
    wb = load_workbook(template_path)
    ws = wb.active

    # Remove example/old data rows, preserve header row 1
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    print(f"🔄 Writing {len(df)} rows...")
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, value in enumerate(map_row(row), start=1):
            ws.cell(row=i, column=col_idx).value = value

    wb.save(output_path)
    print(f"✅ {len(df)} trades written → {output_path}")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='Convert TradeZella CSV export to SmartTraderAI STB format.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument('csv',        help='Path to TradeZella CSV export')
    parser.add_argument('--sheets',   action='store_true', help='Force Google Sheets output')
    parser.add_argument('--xlsx',     action='store_true', help='Force .xlsx file output')
    parser.add_argument('--sheet-id', dest='sheet_id', help='Google Spreadsheet ID (overrides config)')
    parser.add_argument('--creds',    help='Path to service_account.json (overrides config)')
    parser.add_argument('--tab',      default=SHEET_TAB_NAME,
                                      help=f'Sheet tab name (default: {SHEET_TAB_NAME})')
    parser.add_argument('--template', default=TEMPLATE_PATH,
                                      help='Path to STB .xlsx template (xlsx mode only)')
    parser.add_argument('--output',   help='Output .xlsx path (auto-named if omitted)')
    args = parser.parse_args()

    # ── Validate CSV ──────────────────────────────────────────────────────────
    if not os.path.exists(args.csv):
        print(f"❌ CSV not found: {args.csv}")
        sys.exit(1)

    # ── Load + filter CSV ─────────────────────────────────────────────────────
    print(f"📂 Reading: {os.path.basename(args.csv)}")
    df = pd.read_csv(args.csv)
    # Drop any metadata/footer rows that have no Open Date
    df = df[pd.notna(df['Open Date']) & (df['Open Date'].astype(str).str.strip() != '')]
    # Sort chronologically by trade open date
    df['Open Date'] = pd.to_datetime(df['Open Date'])
    df = df.sort_values('Open Date').reset_index(drop=True)
    print(f"   {len(df)} valid trade rows found")

    # ── Resolve credentials ───────────────────────────────────────────────────
    sheet_id = args.sheet_id or SPREADSHEET_ID
    creds    = args.creds    or SERVICE_ACCOUNT_FILE

    # ── Choose output mode ────────────────────────────────────────────────────
    if args.xlsx:
        use_sheets = False
    elif args.sheets:
        use_sheets = True
    else:
        # Auto: use Sheets if both SPREADSHEET_ID and credentials file are set
        configured = (sheet_id != "YOUR_SPREADSHEET_ID_HERE" and os.path.exists(creds))
        use_sheets = configured
        if not configured:
            print("ℹ️  Google Sheets not configured — using .xlsx fallback")
            print("   (See README.md → 'Google Sheets Setup' to enable direct upload)")

    # ── Execute ───────────────────────────────────────────────────────────────
    if use_sheets:
        if sheet_id == "YOUR_SPREADSHEET_ID_HERE":
            print("❌ SPREADSHEET_ID is not set.")
            print("   Edit the SPREADSHEET_ID variable at the top of this script, or use --sheet-id.")
            sys.exit(1)
        if not os.path.exists(creds):
            print(f"❌ service_account.json not found at: {creds}")
            print("   See README.md → 'Google Sheets Setup' for instructions.")
            sys.exit(1)
        write_to_sheets(df, sheet_id, creds, args.tab)

    else:
        if not os.path.exists(args.template):
            print(f"❌ Template not found: {args.template}")
            print("   Place STB_Import_Template.xlsx in the same folder as this script.")
            sys.exit(1)
        script_dir  = os.path.dirname(os.path.abspath(__file__))
        datestamp   = datetime.now().strftime('%Y%m%d')
        output_path = args.output or os.path.join(script_dir, f'STB_Import_Merged_{datestamp}.xlsx')
        write_to_xlsx(df, args.template, output_path)


if __name__ == '__main__':
    main()
